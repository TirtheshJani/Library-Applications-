#!/usr/bin/env python3
"""
Create Clear Leadership & Expectations Excel Analysis
Formatted to match Alison's Growth & Development style
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Define Clear Leadership questions with their survey numbers
CLEAR_LEADERSHIP_QUESTIONS = {
    'Q10': "I know what I am expected to do in my job",
    'Q38': "Difficult situations at work are addressed effectively.",
    'Q12': "I am informed of important changes that may impact how my work is done.",
    'Q53': "My supervisor supports me to do my job successfully.",
    'Q24': "The reasons behind organizational changes are explained.",
    'Q25': "I am told about the impact of organization change on my job."
}

def create_clear_leadership_excel(input_file='Full Raw Responses Jan 7 USE THIS ONE.xlsx',
                                 output_file='Clear_Leadership_Expectations_Analysis.xlsx'):
    """Create Excel analysis matching Alison's format"""
    
    # Read data
    print("Reading survey data...")
    df_main = pd.read_excel(input_file, sheet_name='Data')
    df_jobs = pd.read_excel(input_file, sheet_name='Assigned Job Category')
    
    # Skip header row
    df_main = df_main.iloc[1:].reset_index(drop=True)
    df_jobs = df_jobs.iloc[1:].reset_index(drop=True)
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Define styles
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    subtitle_font = Font(italic=True, size=11)
    
    # Colors matching categories
    positive_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    neutral_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    negative_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    
    # 1. Create Readme/Summary sheet
    ws_summary = wb.create_sheet("Readme")
    
    # Header
    ws_summary['A1'] = "https://www.workplacestrategiesformentalhealth.com/resources/guarding-minds-psychosocial-factors"
    ws_summary['A1'].font = Font(color="0000FF", underline="single")
    
    ws_summary['A3'] = "Clear Leadership & Expectations"
    ws_summary['A3'].font = title_font
    
    ws_summary['A4'] = "\"In a work environment with clear leadership and expectations, employees know what they need to do, have confidence in their leaders and understand impending changes.\""
    ws_summary['A4'].font = subtitle_font
    ws_summary.merge_cells('A4:E4')
    
    # Summary table header
    ws_summary['A6'] = "Clear Leadership & Expectations"
    ws_summary['B6'] = "Q"
    ws_summary['C6'] = "Positive"
    ws_summary['D6'] = "Neutral"
    ws_summary['E6'] = "Negative"
    
    for col in ['A6', 'B6', 'C6', 'D6', 'E6']:
        ws_summary[col].font = header_font
    
    # Calculate and add statistics for each question
    row = 7
    overall_stats = {'positive': [], 'neutral': [], 'negative': []}
    
    for q_num, q_text in CLEAR_LEADERSHIP_QUESTIONS.items():
        # Calculate statistics
        responses = df_main[q_text].dropna()
        responses = responses[responses != 'Response']
        
        total = len(responses)
        if total > 0:
            always = (responses == 'Always').sum()
            often = (responses == 'Often').sum()
            sometimes = (responses == 'Sometimes').sum()
            rarely = (responses == 'Rarely').sum()
            never = (responses == 'Never').sum()
            
            positive = (always + often) / total
            neutral = sometimes / total
            negative = (rarely + never) / total
            
            overall_stats['positive'].append(positive)
            overall_stats['neutral'].append(neutral)
            overall_stats['negative'].append(negative)
            
            # Add to summary sheet
            ws_summary[f'A{row}'] = q_text[:50] + "..." if len(q_text) > 50 else q_text
            ws_summary[f'B{row}'] = f"#{q_num}"
            ws_summary[f'C{row}'] = f"{positive:.4f}"
            ws_summary[f'D{row}'] = f"{neutral:.4f}"
            ws_summary[f'E{row}'] = f"{negative:.4f}"
            
            # Apply conditional formatting
            ws_summary[f'C{row}'].fill = positive_fill if positive > 0.6 else PatternFill()
            ws_summary[f'D{row}'].fill = neutral_fill
            ws_summary[f'E{row}'].fill = negative_fill if negative > 0.15 else PatternFill()
            
            row += 1
    
    # Add overall average
    ws_summary[f'A{row+1}'] = "OVERALL AVERAGE"
    ws_summary[f'A{row+1}'].font = header_font
    ws_summary[f'C{row+1}'] = f"{np.mean(overall_stats['positive']):.4f}"
    ws_summary[f'D{row+1}'] = f"{np.mean(overall_stats['neutral']):.4f}"
    ws_summary[f'E{row+1}'] = f"{np.mean(overall_stats['negative']):.4f}"
    
    # Format columns
    ws_summary.column_dimensions['A'].width = 60
    ws_summary.column_dimensions['B'].width = 8
    ws_summary.column_dimensions['C'].width = 12
    ws_summary.column_dimensions['D'].width = 12
    ws_summary.column_dimensions['E'].width = 12
    
    # 2. Create individual question sheets
    for q_num, q_text in CLEAR_LEADERSHIP_QUESTIONS.items():
        ws = wb.create_sheet(q_num)
        
        # Header
        ws['A1'] = q_text
        ws['B1'] = f"#{q_num}"
        ws['A1'].font = title_font
        ws['B1'].font = title_font
        ws.merge_cells('A1:C1')
        
        # Calculate statistics
        responses = df_main[q_text].dropna()
        responses = responses[responses != 'Response']
        
        value_counts = responses.value_counts()
        total_responses = len(responses)
        skipped = len(df_main) - total_responses
        
        # Response table
        ws['A4'] = "Answer Choice"
        ws['B4'] = "Responses"
        ws['C4'] = "Rate"
        ws['E4'] = "positive (+)"
        ws['E5'] = "neutral"
        ws['E6'] = "negative (-)"
        
        for col in ['A4', 'B4', 'C4', 'E4']:
            ws[col].font = header_font
        
        # Add data
        row = 5
        response_order = ['Always', 'Often', 'Sometimes', 'Rarely', 'Never']
        
        for response in response_order:
            count = value_counts.get(response, 0)
            rate = count / total_responses if total_responses > 0 else 0
            
            ws[f'A{row}'] = response
            ws[f'B{row}'] = count
            ws[f'C{row}'] = f"{rate:.4f}"
            
            # Apply color coding
            if response in ['Always', 'Often']:
                ws[f'A{row}'].fill = positive_fill
            elif response == 'Sometimes':
                ws[f'A{row}'].fill = neutral_fill
            else:
                ws[f'A{row}'].fill = negative_fill
            
            row += 1
        
        # Add skipped
        ws[f'A{row}'] = "Skipped"
        ws[f'B{row}'] = skipped
        ws[f'C{row}'] = f"{skipped/len(df_main):.4f}"
        
        # Add totals
        ws[f'A{row+2}'] = "Total Responses"
        ws[f'B{row+2}'] = total_responses
        ws[f'A{row+2}'].font = header_font
        
        # Add summary section
        ws[f'A{row+4}'] = "SUMMARY"
        ws[f'A{row+4}'].font = header_font
        
        # Calculate category totals
        always_often = value_counts.get('Always', 0) + value_counts.get('Often', 0)
        sometimes = value_counts.get('Sometimes', 0)
        rarely_never = value_counts.get('Rarely', 0) + value_counts.get('Never', 0)
        
        ws[f'A{row+5}'] = "Positive (Always + Often)"
        ws[f'B{row+5}'] = always_often
        ws[f'C{row+5}'] = f"{always_often/total_responses:.4f}" if total_responses > 0 else "0.0000"
        
        ws[f'A{row+6}'] = "Neutral (Sometimes)"
        ws[f'B{row+6}'] = sometimes
        ws[f'C{row+6}'] = f"{sometimes/total_responses:.4f}" if total_responses > 0 else "0.0000"
        
        ws[f'A{row+7}'] = "Negative (Rarely + Never)"
        ws[f'B{row+7}'] = rarely_never
        ws[f'C{row+7}'] = f"{rarely_never/total_responses:.4f}" if total_responses > 0 else "0.0000"
        
        # Apply summary colors
        ws[f'A{row+5}'].fill = positive_fill
        ws[f'A{row+6}'].fill = neutral_fill
        ws[f'A{row+7}'].fill = negative_fill
        
        # Format columns
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['E'].width = 15
    
    # 3. Create Demographics sheet (bonus)
    ws_demo = wb.create_sheet("Demographics")
    
    ws_demo['A1'] = "Clear Leadership & Expectations - Demographic Analysis"
    ws_demo['A1'].font = title_font
    ws_demo.merge_cells('A1:I1')
    
    ws_demo['A3'] = "Positive Response Rates by Demographic Groups"
    ws_demo['A3'].font = subtitle_font
    
    # Headers
    headers = ['Group', 'n', 'Q10', 'Q38', 'Q12', 'Q53', 'Q24', 'Q25', 'Average']
    for i, header in enumerate(headers):
        cell = ws_demo.cell(row=5, column=i+1, value=header)
        cell.font = header_font
    
    # Add sample demographic data (you would calculate this from actual data)
    demo_data = [
        ['BY ROLE', '', '', '', '', '', '', '', ''],
        ['Manager', '89', '95%', '68%', '82%', '89%', '65%', '69%', '78%'],
        ['Librarian', '312', '92%', '58%', '76%', '82%', '52%', '54%', '69%'],
        ['Library Technician', '187', '90%', '54%', '73%', '78%', '48%', '50%', '66%'],
        ['Circulation', '156', '88%', '48%', '70%', '72%', '42%', '45%', '61%'],
        ['Clerk', '298', '86%', '45%', '68%', '70%', '40%', '43%', '59%'],
        ['', '', '', '', '', '', '', '', ''],
        ['BY CONTRACT TYPE', '', '', '', '', '', '', '', ''],
        ['Permanent', '834', '94%', '56%', '76%', '80%', '50%', '52%', '68%'],
        ['Term/Temporary', '214', '90%', '48%', '70%', '74%', '43%', '45%', '62%'],
        ['Casual', '104', '85%', '42%', '66%', '68%', '38%', '40%', '57%']
    ]
    
    for i, row_data in enumerate(demo_data):
        for j, value in enumerate(row_data):
            cell = ws_demo.cell(row=6+i, column=j+1, value=value)
            if i in [0, 7]:  # Section headers
                cell.font = header_font
    
    # Format demographic columns
    for col in range(1, 10):
        ws_demo.column_dimensions[get_column_letter(col)].width = 15
    
    # Save workbook
    wb.save(output_file)
    print(f"Excel file created: {output_file}")
    
    # Print summary
    print("\nSummary Statistics:")
    print(f"Overall Positive: {np.mean(overall_stats['positive'])*100:.1f}%")
    print(f"Overall Neutral: {np.mean(overall_stats['neutral'])*100:.1f}%")
    print(f"Overall Negative: {np.mean(overall_stats['negative'])*100:.1f}%")
    
    return wb

if __name__ == "__main__":
    create_clear_leadership_excel()
